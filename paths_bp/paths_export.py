"""
export results in XLSX format
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from tempfile import NamedTemporaryFile

def export_paths(paths):

    wb = Workbook()

    ws1 = wb.active
    ws1.title = f"Paths"

    header = ["WA code", "Sample ID", "Date", "Municipality", "Coordinates WGS84 UTM zone 32N", "mtDNA result",
                "Genotype ID", "Temp ID", "Sex", "Status", "Pack", "Dead recovery"]
    for locus in loci_list:
        header.extend([f"{locus} a", f"Notes for {locus} a"])
        if loci_list[locus] == 2:
            header.extend([f"{locus} b", f"Notes for {locus} b"])

    ws1.append(header)

    for row in wa_scats:
        out = []
        out.append(row["wa_code"])
        out.append(row["sample_id"] if row["sample_id"] is not None else "")
        out.append(row["date"] if row["date"] is not None else "")
        out.append(row["municipality"] if row["municipality"] is not None else "")
        out.append(f'{row["coord_east"]}, {row["coord_north"]}')
        out.append(row["mtdna"] if row["mtdna"] is not None else "")
        out.append(row["genotype_id"] if row["genotype_id"] is not None else "")
        out.append(row["tmp_id"] if row["tmp_id"] is not None else "")
        out.append(row["sex_id"] if row["sex_id"] is not None else "")
        out.append(row["status"] if row["status"] is not None else "")
        out.append(row["pack"] if row["pack"] is not None else "")
        out.append(row["dead_recovery"] if row["dead_recovery"] is not None else "")

        for locus in loci_list:
            out.extend([loci_values[row['wa_code']][locus]['a']['value'],
                        loci_values[row['wa_code']][locus]['a']['notes'],
                        loci_values[row['wa_code']][locus]['b']['value'],
                        loci_values[row['wa_code']][locus]['b']['notes'],
                       ]
                      )

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream


