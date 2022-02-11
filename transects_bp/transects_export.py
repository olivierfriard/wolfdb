"""
export results in XLSX format
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from tempfile import NamedTemporaryFile

def export_transects(paths):

    wb = Workbook()

    ws1 = wb.active
    ws1.title = f"Transects"

    header = ["Transect ID", "Sector", "Location", "Municipality", "Province", "Region"]

    ws1.append(header)

    for row in paths:
        out = []
        out.append(row["transect_id"])
        out.append(row["sector"])
        out.append(row["location"])
        out.append(row["municipality"])
        out.append(row["province"])
        out.append(row["region"])

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream


