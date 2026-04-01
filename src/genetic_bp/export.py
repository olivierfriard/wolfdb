"""
WolfDB web service
(c) Olivier Friard

export results in XLSX format
"""

from io import BytesIO
from tempfile import NamedTemporaryFile

import pandas as pd
from openpyxl import Workbook


def export_wa_genetic_samples(loci_list: dict, wa_scats, loci_values, file_format: str):
    """
    export genetic data for wa codes
    """

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "WA genetic samples"

    fields = {
        "wa_code": "WA code",
        "sample_id": "Sample ID",
        "date": "Date",
        "municipality": "Municipality",
        "province": "Province",
        "coord_east": "Coordinate WGS84 UTM East",
        "coord_north": "Coordinate WGS84 UTM North",
        "coord_zone": "UTM Zone",
        "mtdna": "mtDNA result",
        "genotype_id": "Genotype ID",
        "notes": "Notes on genotype",
        "tmp_id": "Temp ID",
        "sex_id": "Sex",
        "status": "Status",
        "pack": "Pack",
        "dead_recovery": "Dead recovery",
    }

    header: list = list(fields.values())

    for locus in loci_list:
        for allele in ("a", "b")[: loci_list[locus]]:
            header.extend([f"{locus} {allele}", f"Notes for {locus} {allele}"])

    ws1.append(header)

    for row in wa_scats:
        out: list = [row[field] if row[field] is not None else "" for field in fields]

        for locus in loci_list:
            for allele in ("a", "b")[: loci_list[locus]]:
                out.extend(
                    [
                        loci_values[row["wa_code"]][locus][allele]["value"],
                        loci_values[row["wa_code"]][locus][allele]["notes"],
                    ]
                )

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream


def export_wa_genetic_samples_pandas(
    loci_list: dict,
    wa_scats,
    loci_values,
    file_format: str,
):
    """
    Export genetic data for WA codes.

    Supported formats:
    - xlsx
    - ods
    - tsv

    Returns:
        bytes: file content
    """

    fields = {
        "wa_code": "WA code",
        "sample_id": "Sample ID",
        "date": "Date",
        "municipality": "Municipality",
        "province": "Province",
        "coord_east": "Coordinate WGS84 UTM East",
        "coord_north": "Coordinate WGS84 UTM North",
        "coord_zone": "UTM Zone",
        "mtdna": "mtDNA result",
        "genotype_id": "Genotype ID",
        "notes": "Notes on genotype",
        "tmp_id": "Temp ID",
        "sex_id": "Sex",
        "status": "Status",
        "pack": "Pack",
        "dead_recovery": "Dead recovery",
    }

    rows = []

    for row in wa_scats:
        out = {}

        for field_key, field_label in fields.items():
            value = row[field_key]
            out[field_label] = value if value is not None else ""

        wa_code = row["wa_code"]

        for locus in loci_list:
            for allele in ("a", "b")[: loci_list[locus]]:
                locus_data = loci_values.get(wa_code, {}).get(locus, {}).get(allele, {})

                out[f"{locus} {allele}"] = locus_data.get("value", "")
                out[f"Notes for {locus} {allele}"] = locus_data.get("notes", "")

        rows.append(out)

    df = pd.DataFrame(rows)

    file_format = file_format.lower().strip()

    if file_format == "xlsx":
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="WA genetic samples", index=False)
        output.seek(0)
        return output.getvalue()

    elif file_format == "ods":
        with NamedTemporaryFile(suffix=".ods") as tmp:
            with pd.ExcelWriter(tmp.name, engine="odf") as writer:
                df.to_excel(writer, sheet_name="WA genetic samples", index=False)
            tmp.seek(0)
            return tmp.read()

    elif file_format == "tsv":
        output = BytesIO()
        df.to_csv(output, sep="\t", index=False, encoding="utf-8")
        output.seek(0)
        return output.getvalue()

    else:
        raise ValueError(f"Unsupported file format: {file_format}")


def export_wa_analysis(
    loci_list, wa_scats, loci_values, distance: int, cluster_id: int
):
    """
    export cluster content in XLSX
    """

    wb = Workbook()

    ws1 = wb.active
    ws1.title = f"WA matches (DBSCAN distance {distance} cluster ID {cluster_id})"

    fields = {
        "wa_code": "WA code",
        "sample_id": "Sample ID",
        "date": "Date",
        "municipality": "Municipality",
        "coord_east": "Coordinate WGS84 UTM East",
        "coord_north": "Coordinate WGS84 UTM North",
        "coord_zone": "UTM Zone",
        "mtdna": "mtDNA result",
        "genotype_id": "Genotype ID",
        "notes": "Notes on genotype",
        "tmp_id": "Temporary ID",
        "sex_id": "Sex",
        "status": "Status",
        "pack": "Pack",
        "dead_recovery": "Dead recovery",
    }

    header: list = list(fields.values())
    for locus in loci_list:
        header.extend([f"{locus} a", f"Notes for {locus} a"])
        if loci_list[locus] == 2:
            header.extend([f"{locus} b", f"Notes for {locus} b"])

    ws1.append(header)

    for row in wa_scats:
        out: list = [row[field] if row[field] is not None else "" for field in fields]
        """
        for locus in loci_list:
            out.extend(
                [
                    loci_values[row["wa_code"]][locus]["a"]["value"],
                    loci_values[row["wa_code"]][locus]["a"]["notes"],
                    loci_values[row["wa_code"]][locus]["b"]["value"],
                    loci_values[row["wa_code"]][locus]["b"]["notes"],
                ]
            )
        """

        for locus in loci_list:
            for allele in ("a", "b")[: loci_list[locus]]:
                out.extend(
                    [
                        loci_values[row["wa_code"]][locus][allele]["value"],
                        loci_values[row["wa_code"]][locus][allele]["notes"],
                    ]
                )

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream


def export_wa_analysis_group(loci_list, data, loci_values):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Genotype matches"

    fields = {
        "working_notes": "Notes on genotype",
        "tmp_id": "Temporary ID",
        "sex": "Sex",
        "status": "Status",
        "pack": "Pack",
        "hybrid": "Hybrid",
        "dispersal": "Dispersal",
        "n_recap": "Number of recaptures",
        "dead_recovery": "Dead recovery",
    }

    header = ["Genotype ID"] + list(fields.values())

    for locus in loci_list:
        header.extend([f"{locus} a", f"Notes for {locus} a"])
        if loci_list[locus] == 2:
            header.extend([f"{locus} b", f"Notes for {locus} b"])

    ws1.append(header)

    for genotype_id in data:
        out = [genotype_id]

        for field in fields:
            out.append(data[genotype_id][field])

        """for locus in loci_list:
            out.extend(
                [
                    loci_values[genotype_id][locus]["a"]["value"],
                    loci_values[genotype_id][locus]["a"]["notes"],
                    loci_values[genotype_id][locus]["b"]["value"] if "b" in loci_values[genotype_id][locus] else "",
                    loci_values[genotype_id][locus]["b"]["notes"] if "b" in loci_values[genotype_id][locus] else "",
                ]
            )
        """

        for locus in loci_list:
            for allele in ("a", "b")[: loci_list[locus]]:
                out.extend(
                    [
                        loci_values[genotype_id][locus][allele]["value"],
                        loci_values[genotype_id][locus][allele]["notes"],
                    ]
                )

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream


def export_wa(loci_list, wa_list, loci_values):
    """
    export WA in XLSX
    """

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "WA"

    fields = {
        "wa_code": "WA code",
        "sample_id": "Sample ID",
        "genotype_id": "Genotype ID",
        "date": "Date",
        "box_number": "Box number",
        "municipality": "Municipality",
        "province": "Province",
        "coord_east": "Coordinates East WGS84 UTM",
        "coord_north": "Coordinates North WGS84 UTM",
        "coord_zone": "UTM Zone",
        "mtdna": "mtDNA result",
        "tmp_id": "Temporary ID",
        "sex_id": "Sex",
        "status": "Status",
        "pack": "Pack",
        "dead_recovery": "Dead recovery",
    }

    header: list = list(fields.values())

    for locus in loci_list:
        header.extend([f"{locus} a", f"Notes for {locus} a"])
        if loci_list[locus] == 2:
            header.extend([f"{locus} b", f"Notes for {locus} b"])

    ws1.append(header)

    for wa in wa_list:
        out: list = [wa[field] if wa[field] is not None else "" for field in fields]

        for locus in loci_list:
            if wa["wa_code"] in loci_values:
                out.extend(
                    [
                        loci_values[wa["wa_code"]][locus]["a"]["value"],
                        loci_values[wa["wa_code"]][locus]["a"]["notes"],
                        loci_values[wa["wa_code"]][locus]["b"]["value"]
                        if "b" in loci_values[wa["wa_code"]][locus]
                        else "",
                        loci_values[wa["wa_code"]][locus]["b"]["notes"]
                        if "b" in loci_values[wa["wa_code"]][locus]
                        else "",
                    ]
                )
            else:
                out.extend(["", "", "", ""])

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream


def export_genotypes_list(loci_list, results, loci_values):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Genotype matches"

    fields = {
        "genotype_id": "Genotype ID",
        "tmp_id": "Other ID",
        "date": "Date",
        "pack": "Pack",
        "sex": "Sex",
        "hybrid": "Hybrid",
        "status": "Status",
        "age_first_capture": "Age at first capture",
        "status_first_capture": "status at first capture",
        "dispersal": "Dispersal",
        "n_recaptures": "Number of recaptures",
        "dead_recovery": "Dead recovery",
    }

    header: list = list(fields.values())
    for locus in loci_list:
        for allele in ("a", "b")[: loci_list[locus]]:
            header.extend([f"{locus} {allele}", f"Notes for {locus} {allele}"])

    ws1.append(header)

    for row in results:
        out: list = [row[field] if row[field] is not None else "" for field in fields]
        for locus in loci_list:
            for allele in ("a", "b")[: loci_list[locus]]:
                out.extend(
                    [
                        loci_values[row["genotype_id"]][locus][allele]["value"],
                        loci_values[row["genotype_id"]][locus][allele]["notes"],
                    ]
                )

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream
