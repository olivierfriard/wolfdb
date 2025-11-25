"""
export paths in XLSX format
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from tempfile import NamedTemporaryFile


def export_paths(paths):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = f"Paths"

    header = [
        "Path ID",
        "Transect ID",
        "Region",
        "Province",
        "Date",
        "Sampling season",
        "Completeness",
        "Number of samples",
        "Number of tracks",
        "Operator",
        "Institution",
        "Category",
        "Notes",
    ]

    ws1.append(header)

    for row in paths:
        out = []
        out.append(row["path_id"])
        out.append(row["transect_id"])
        out.append(row["region"])
        out.append(row["province"])
        out.append(row["date"] if row["date"] is not None else "")
        out.append(row["sampling_season"] if row["sampling_season"] is not None else "")
        out.append(row["completeness"])
        out.append(row["n_samples"])
        out.append(row["n_tracks"])
        out.append(row["observer"] if row["observer"] is not None else "")
        out.append(row["institution"] if row["institution"] is not None else "")
        out.append(row["category"] if row["category"] is not None else "")
        out.append(row["notes"] if row["notes"] is not None else "")

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream
