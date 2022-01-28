"""
export results in XLSX format
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from tempfile import NamedTemporaryFile

def export_wa_genetic_samples(loci_list, wa_scats, loci_values, with_notes):

    wb = Workbook()

    ws1 = wb.active
    ws1.title = f"WA genetic samples"

    header = ["WA code", "Sample ID", "Date", "Municipality", "Coordinates WGS84 UTM zone 32N", "mtDNA result",
                "Genotype ID", "Temp ID", "Sex", "Status", "Pack", "Dead recovery"]
    for locus in loci_list:
        header.extend([f"{locus} a", f"Notes for {locus} a"])
        if loci_list[locus] == 2:
            header.extend([f"{locus} b", f"Notes for {locus} b"])

    ws1.append(header)

    for row in wa_scats:
        out = []
        out.append(row["wa_code"])
        out.append(row["sample_id"] if row["sample_id"] is not None else "")
        out.append(row["date"] if row["date"] is not None else "")
        out.append(row["municipality"] if row["municipality"] is not None else "")
        out.append(f'{row["coord_east"]}, {row["coord_north"]}')
        out.append(row["mtdna"] if row["mtdna"] is not None else "")
        out.append(row["genotype_id"] if row["genotype_id"] is not None else "")
        out.append(row["tmp_id"] if row["tmp_id"] is not None else "")
        out.append(row["sex_id"] if row["sex_id"] is not None else "")
        out.append(row["status"] if row["status"] is not None else "")
        out.append(row["pack"] if row["pack"] is not None else "")
        out.append(row["dead_recovery"] if row["dead_recovery"] is not None else "")

        for locus in loci_list:
            out.extend([loci_values[row['wa_code']][locus]['a']['value'],
                        loci_values[row['wa_code']][locus]['a']['notes'],
                        loci_values[row['wa_code']][locus]['b']['value'],
                        loci_values[row['wa_code']][locus]['b']['notes'],
                       ]
                      )

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream


def export_wa_analysis(loci_list, wa_scats, loci_values, distance, cluster_id):

    wb = Workbook()

    ws1 = wb.active
    ws1.title = f"WA matches (DBSCAN distance {distance} cluster ID {cluster_id})"
    print(ws1.title)

    header = ["WA code", "Sample ID", "Date", "Municipality", "Coordinates WGS84 UTM zone 32N", "mtDNA result",
              "Genotype ID", "Temporary ID", "Sex", "Status", "Pack", "Dead recovery"]
    for locus in loci_list:
        header.extend([f"{locus} a", f"Notes for {locus} a"])
        if loci_list[locus] == 2:
            header.extend([f"{locus} b", f"Notes for {locus} b"])

    ws1.append(header)

    for row in wa_scats:
        out = []
        out.append(row["wa_code"])
        out.append(row["sample_id"] if row["sample_id"] is not None else "")
        out.append(row["date"] if row["date"] is not None else "")
        out.append(row["municipality"] if row["municipality"] is not None else "")
        out.append(f'{row["coord_east"]}, {row["coord_north"]}')
        out.append(row["mtdna"] if row["mtdna"] is not None else "")
        out.append(row["genotype_id"] if row["genotype_id"] is not None else "")
        out.append(row["tmp_id"] if row["tmp_id"] is not None else "")
        out.append(row["sex_id"] if row["sex_id"] is not None else "")

        out.append(row["status"] if row["status"] is not None else "")
        out.append(row["pack"] if row["pack"] is not None else "")
        out.append(row["dead_recovery"] if row["dead_recovery"] is not None else "")



        for locus in loci_list:
            out.extend([loci_values[row['wa_code']][locus]['a']['value'],
                        loci_values[row['wa_code']][locus]['a']['notes'],
                        loci_values[row['wa_code']][locus]['b']['value'],
                        loci_values[row['wa_code']][locus]['b']['notes'],
                       ]
                      )

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream



def export_wa_analysis_group(loci_list, data, loci_values):

    wb = Workbook()

    ws1 = wb.active
    ws1.title = f"Genotype matches"

    header = ["Genotype ID", "Temporary ID", "Sex", "Status", "Pack", "Number of recaptures"]
    for locus in loci_list:
        header.extend([f"{locus} a", f"Notes for {locus} a"])
        if loci_list[locus] == 2:
            header.extend([f"{locus} b", f"Notes for {locus} b"])

    ws1.append(header)

    for genotype_id in data:
        out = []
        
        out.append(genotype_id)
        out.append(data[genotype_id]["tmp_id"])
        out.append(data[genotype_id]["sex"])
        out.append(data[genotype_id]["position"])
        out.append(data[genotype_id]["pack"])
        out.append(data[genotype_id]["n_recap"])

        for locus in loci_list:
            out.extend([loci_values[genotype_id][locus]['a']['value'],
                        loci_values[genotype_id][locus]['a']['notes'],
                        loci_values[genotype_id][locus]['b']['value'],
                        loci_values[genotype_id][locus]['b']['notes'],
                       ]
                      )

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream


def export_genotypes_list(loci_list, results, loci_values):

    wb = Workbook()

    ws1 = wb.active
    ws1.title = f"Genotype matches"

    header = ["Genotype ID", "Other ID", "Date", "Pack", "Sex", "Age at first capture", "status at first capture",
              "Dispersal", "Number of recaptures", "Dead recovery"]
    for locus in loci_list:
        header.extend([f"{locus} a", f"Notes for {locus} a"])
        if loci_list[locus] == 2:
            header.extend([f"{locus} b", f"Notes for {locus} b"])

    ws1.append(header)

    for row in results:
        out = []
        out.append(row["genotype_id"])
        out.append(row["tmp_id"] if row["tmp_id"] is not None else "")
        out.append(row["date"] if row["date"] is not None else "")
        out.append(row["pack"] if row["pack"] is not None else "")
        out.append(row["sex"] if row["sex"] is not None else "")
        out.append(row["age_first_capture"] if row["age_first_capture"] is not None else "")
        out.append(row["status_first_capture"] if row["status_first_capture"] is not None else "")

        out.append(row["dispersal"] if row["dispersal"] is not None else "")
        out.append(row["n_recaptures"] if row["n_recaptures"] is not None else "")
        out.append(row["dead_recovery"] if row["dead_recovery"] is not None else "")        

        for locus in loci_list:
            out.extend([loci_values[row['genotype_id']][locus]['a']['value'],
                        loci_values[row['genotype_id']][locus]['a']['notes'],
                        loci_values[row['genotype_id']][locus]['b']['value'],
                        loci_values[row['genotype_id']][locus]['b']['notes'],
                       ]
                      )

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream