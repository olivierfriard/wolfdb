"""
export scats in XLSX format
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from tempfile import NamedTemporaryFile


def export_scats(scats):

    wb = Workbook()

    ws1 = wb.active
    ws1.title = f"Scats"

    header:list = [
        "Scat ID",
        "Date",
        "Sampling season",
        "Sampling type",
        "path ID",
        "Snowtrack ID",
        "WA code",
        "Genotype ID",
        "location",
        "municipality",
        "province",
        "region",
        "Deposition",
        "Matrix",
        "Collected scat",
        "Scalp category",
        "Genetic sample",
        "Coordinate east",
        "Coordinate north",
        "Zone",
        "Operator",
        "Institution",
        "Notes",
    ]

    ws1.append(header)

    for row in scats:
        out = []
        out.append(row["scat_id"])
        out.append(row["date"])
        out.append(row["sampling_season"])
        out.append(row["sampling_type"])
        out.append(row["path_id"] if row["path_id"] is not None else "")
        out.append(row["snowtrack_id"] if row["snowtrack_id"] is not None else "")
        out.append(row["wa_code"] if row["wa_code"] is not None else "")
        out.append(row["genotype_id2"])
        out.append(row["location"] if row["location"] is not None else "")
        out.append(row["municipality"] if row["municipality"] is not None else "")
        out.append(row["province"] if row["province"] is not None else "")
        out.append(row["region"] if row["region"] is not None else "")
        out.append(row["deposition"])
        out.append(row["matrix"])
        out.append(row["collected_scat"])
        out.append(row["scalp_category"])

        """
        scalp_cat = row["scalp_category"]
        if row["mtdna"] is not None and "WOLF" in row["mtdna"].upper():
            scalp_cat += " (from mtDNA: C1)"
        out.append(scalp_cat)
        """

        out.append(row["genetic_sample"])
        out.append(row["coord_east"])
        out.append(row["coord_north"])
        out.append(row["coord_zone"])
        out.append(row["observer"])
        out.append(row["institution"])
        out.append(row["notes"])

        ws1.append(out)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

        return stream
